import sqlite3
from openpyxl import load_workbook
import os
from pathlib import Path

# print(os.path.expanduser("~"))
# Get the current working directory
current_dir = os.getcwd()

# Navigate two steps back from the parent directory of the current working directory
two_steps_back_dir = os.path.dirname(current_dir)
# Define the Excel file path within the two-steps-back directory
excel_file_path = os.path.join(two_steps_back_dir, "2wdata.xlsx")
db_file_path = os.path.join(two_steps_back_dir, "projects", "execution_folder", "config_folder","config_file_abc9003.db")

# # Load the Excel workbook and select the worksheet
workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active


# Connect to the SQLite database or create it if it doesn't exist
conn = sqlite3.connect(db_file_path)

cursor = conn.cursor()

# Create the 'testdatatable' table if it doesn't exist
cursor.execute('''
    CREATE TABLE IF NOT EXISTS testdatatable (
        step INTEGER,
        var TEXT,
        actual_data TEXT
    )
''')

# Define the range in Excel to read from
start_row = input("enter start row")  # B4 in Excel
end_row = input("enter end Row")   # B17 in Excel
values=[]
var=None
step=None
# Loop through the rows in Excel and insert data into the database
for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
    actual_data=""
    for value in row[1:]:
        if value:
              actual_data+=str(value)+",,"
        else:
            break
    actual_data=actual_data[:-2]
    values.append(actual_data) # Assuming 'some_var_value' for var column

     # Check if the data already exists in the database based on 'step' value
    cursor.execute('''
        SELECT * FROM testdatatable WHERE actual_data = ?
    ''', (actual_data,))
    
    existing_data = cursor.fetchone()
    print(existing_data)
    if existing_data:
        # Data already exists, update the existing row
        cursor.execute('''
            UPDATE testdatatable 
            SET var = ?, step = ?
            WHERE actual_data = ?
        ''', (var,step, actual_data))
    else:
        # Data doesn't exist, insert a new row
        cursor.execute('''
            INSERT INTO testdatatable (step, var, actual_data) VALUES (?, ?, ?)
        ''', (step, var, actual_data))

print(values)

# Commit the changes and close the database connection
conn.commit()
conn.close()

print("Data inserted into SQLite database.")
