from openpyxl import Workbook
from openpyxl import load_workbook
import os
import random
current_dir = os.getcwd()

two_steps_back_dir = os.path.dirname(current_dir)
excel_file_path = os.path.join(two_steps_back_dir, "dta.xlsx")

dic={
    "channel":["agent","broker"],
    "proposalno":["2w0010","2w0011","2w0010"],
    "salutation":["Mr.","Mrs."],
    "dob":["15061990","1752006","222000"],
    "gender":["Male","Female"],
    "annutype":["1.1","1.2","1.3"],
    "defermentperiod":["1","2","3","4"],
    "subplan":["2.1","2.2"],
    "annuitypayoutfreq":["Monthly","Quarterly"],
    "purchaseprice":["1999","2999"],
    "annuitypurchaseprice":["4999","9999"],
    "joint":["Yes","No"],
    "sys_id":["789","456","123","000"],
    "Product name":["SBI Annuity"],
    "Proposal Category":["Individual"],
    "SA":[100,200],
    "Frequency":["monthly","yearly"],
    "proposalno":[10,9,8],
    "Country":["India","Japan"],
    "Nationality":["India"],
    "AgeProofDocument":["Aadhar"],
    "Hazard":["Yes","No"],
    "SBI":["Yes"],
    "Deferment period":["1,2"],
    "HEIGHT":[180,170],
    "WEIGHT":[50,60],
    "BP":["Yes","No"],
    "Smoking":["Yes","No"],
    "Proofofidentity":["Aadhar"],
    "MaritalStatus":['Married',"single"],
    "Resident":["India"],
    "PostalCode":[524001,420421,423423],
    

}

workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active
def generate_columns(dropdown_selection, startRow, endRow,num_test_cases):

    sr=startRow
    for row in worksheet.iter_rows(min_row=startRow, max_row=endRow, values_only=True):
        print(sr,row)
        if row[0] in dic:
            value=dic[row[0]]
            for i in range(2,num_test_cases + 2):
                worksheet.cell(row=sr, column=i, value=random.choice(value))
        sr+=1

    # Save the workbook
    workbook.save(excel_file_path)


data=worksheet["A2"]
num_test_cases=worksheet["A3"]
if data.value=="2w":
    startRow=4
    endRow=22
else:
    startRow=26
    endRow=69

# Example usage
generate_columns(data.value,startRow,endRow, num_test_cases.value)
