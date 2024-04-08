import sqlite3
from openpyxl import load_workbook
import os

current_dir = os.getcwd()

two_steps_back_dir = os.path.dirname(current_dir)
excel_file_path = os.path.join(two_steps_back_dir, "dta.xlsx")
db_file_path = os.path.join(two_steps_back_dir, "projects", "execution_folder", "config_folder", "config_file_abc9003.db")

workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active

conn = sqlite3.connect(db_file_path)
cursor = conn.cursor()

data = worksheet["A2"]
num_test_cases = worksheet["A3"]
if data.value == "2w":
    start_row = 4
    end_row = 22
else:
    start_row = 26
    end_row = 69

values = []
# Loop through the rows in Excel and insert data into the database
for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
    title = row[0]  # Assuming the title is in the first column of each row
    data = ""
    for value in row[1:]:
        if value:
            data += str(value) + ",,"
        else:
            break
    data = data[:-2]

    # Check if the data already exists in the database
    cursor.execute('''
        SELECT * FROM testdatatable WHERE title = ?
    ''', (title,))

    existing_data = cursor.fetchone()
    print(existing_data)

    if existing_data and data:
        cursor.execute('''
            UPDATE testdatatable 
            SET data=?
            WHERE title = ?
        ''', (data, title))

conn.commit()
conn.close()

print("Data inserted/updated in SQLite database.")

